"""Pipeline Descentralizados — Operativa Futuros Pionex.

Lee los exports crudos de Pionex desde data/:
  - position_futures.csv      (posiciones cerradas: pnl, fee, funding)
  - raw-trading-details.csv   (fills: para contar promedios y capital desplegado)
  - others.csv                (RiskCoverage: para detectar liquidaciones)
  - operaciones-manuales.csv  (opcional: operaciones aún no incluidas en el export)

Genera:
  - dist/index.html                      (dashboard)
  - dist/operativa-futuros-pionex.xlsx   (Excel de reporte, descargable desde el dashboard)

Para actualizar la operativa basta con reexportar los CSV de Pionex y
sobrescribirlos en data/. Las operaciones manuales se deduplican solas cuando
el export ya las incluye (mismo símbolo y cierre a menos de 2 minutos).
"""
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DIST = ROOT / "dist"
TEMPLATE = ROOT / "templates" / "dashboard.template.html"


# ---------------------------------------------------------------- carga

def load_inputs():
    pos = pd.read_csv(DATA / "position_futures.csv", parse_dates=["open_time", "close_time"])
    raw = pd.read_csv(DATA / "raw-trading-details.csv", parse_dates=["date(UTC+0)"])
    raw = raw[(raw.market_type == "Futures USDT") & (raw.state == "FILLED")].rename(
        columns={"date(UTC+0)": "ts"}
    )
    raw["subid"] = raw.tax_id.str.replace("s_", "", regex=False)
    oth = pd.read_csv(DATA / "others.csv", parse_dates=["date(UTC+0)"])
    risk = oth[oth.tag == "RiskCoverage"].rename(columns={"date(UTC+0)": "ts"})
    return pos, raw, risk


# ------------------------------------------------- asignación de fills

def assign_fills(pos, raw):
    """Asigna cada fill a una única posición (subid + ventana temporal)."""
    T3 = pd.Timedelta(seconds=3)
    assign = {}
    for subid, g in pos.groupby("subid"):
        g = g.sort_values("open_time")
        fills = raw[raw.subid == subid].sort_values("ts")
        for fi, r in fills.iterrows():
            cands = g[(g.open_time - T3 <= r.ts) & (g.close_time + T3 >= r.ts)]
            if len(cands) == 1:
                assign[fi] = cands.iloc[0].pid
            elif len(cands) > 1:
                # conflicto de frontera entre posiciones consecutivas
                if (r.side == "SELL") == (cands.iloc[0].position_side == "short"):
                    near = cands.iloc[(cands.open_time - r.ts).abs().argmin()]
                else:
                    near = cands.iloc[(cands.close_time - r.ts).abs().argmin()]
                assign[fi] = near.pid
            else:
                # fill de apertura registrado antes del open_time oficial
                g2 = g[g.close_time + T3 >= r.ts]
                if len(g2):
                    cand = g2.iloc[0]
                    prev = g[g.close_time < cand.open_time]
                    lo = prev.close_time.max() if len(prev) else pd.Timestamp("2000-01-01")
                    if r.ts > lo:
                        assign[fi] = cand.pid
    return assign


def count_entry_events(opens):
    """Agrupa fills de entrada en eventos (gap>10min o precio >1% distinto)."""
    ev, lts, lpx = 0, None, None
    for _, r in opens.iterrows():
        if lts is None or (r.ts - lts) > pd.Timedelta(minutes=10) or (
            lpx and abs(r.price - lpx) / lpx > 0.01
        ):
            ev += 1
        lts, lpx = r.ts, r.price
    return ev


# ------------------------------------------------------------ análisis

def analyze(pos, raw, risk):
    pos = pos.sort_values("open_time").reset_index(drop=True)
    pos["pid"] = pos.index
    pos["subid"] = pos.position_id.str.split("_").str[1]
    pos["is_bot"] = pos.subid.str.len() < 10

    raw = raw.copy()
    raw["pid"] = raw.index.map(assign_fills(pos, raw))

    rows = []
    for _, p in pos.iterrows():
        f = raw[raw.pid == p.pid].sort_values("ts")
        if p.position_side == "short":
            opens = f[f.side == "SELL"]
        else:
            opens = f[f.side == "BUY"]
        entries = count_entry_events(opens)
        liq = (
            (risk.ts >= p.close_time - pd.Timedelta(seconds=15))
            & (risk.ts <= p.close_time + pd.Timedelta(seconds=15))
        ).any()
        rows.append(
            dict(
                symbol=p.symbol.replace("_USDT_PERP", ""),
                side=p.position_side,
                open=p.open_time,
                close=p.close_time,
                pnl=p.pnl,
                fee=p.fee,
                funding=p.funding_fee,
                promedios=max(entries - 1, 0),
                capital=opens.amount.sum(),
                liq=bool(liq),
                is_bot=bool(p.is_bot),
            )
        )
    return pd.DataFrame(rows)


def load_manual_ops(df):
    """Añade operaciones manuales aún no presentes en el export (dedup ±2 min)."""
    path = DATA / "operaciones-manuales.csv"
    if not path.exists():
        return df
    man = pd.read_csv(path, parse_dates=["open_time", "close_time"])
    added = []
    for _, r in man.iterrows():
        sym = str(r.symbol).replace("_USDT_PERP", "")
        dup = df[
            (df.symbol == sym)
            & ((df.close - r.close_time).abs() < pd.Timedelta(minutes=2))
        ]
        if len(dup):
            continue  # el export ya la incluye
        added.append(
            dict(
                symbol=sym,
                side=str(r.side).lower(),
                open=r.open_time,
                close=r.close_time,
                pnl=float(r.pnl),
                fee=float(r.fee) if pd.notna(r.fee) else 0.0,
                funding=float(r.funding) if pd.notna(r.funding) else 0.0,
                promedios=int(r.promedios) if pd.notna(r.promedios) else 0,
                capital=float(r.capital),
                liq=str(r.liquidada).strip().lower() in ("si", "sí", "yes", "true", "1"),
                is_bot=False,
            )
        )
    if added:
        df = pd.concat([df, pd.DataFrame(added)], ignore_index=True)
    return df


def finalize(df):
    df = df.sort_values("close").reset_index(drop=True)
    df["net"] = df.pnl + df.fee + df.funding
    df["tipo"] = np.where(df.is_bot, "Bot", "Manual")

    def resultado(r):
        if r.liq:
            return "Liquidada"
        if r.net > 0:
            return "TP directo" if r.promedios == 0 else f"TP tras {int(r.promedios)} promedio(s)"
        return "Cerrada en pérdida"

    df["resultado"] = df.apply(resultado, axis=1)
    df["rent"] = np.where(df.capital > 0, df.net / df.capital, np.nan)
    return df


# -------------------------------------------------------------- salidas

def write_html(df):
    ops = []
    for _, r in df.iterrows():
        ops.append(
            {
                "d": r.close.strftime("%Y-%m-%d %H:%M"),
                "s": r.symbol,
                "l": "Short" if r.side == "short" else "Long",
                "t": r.tipo,
                "p": int(r.promedios),
                "c": round(float(r.capital), 2),
                "n": round(float(r.net), 2),
                "r": round(float(r.rent) * 100, 2) if pd.notna(r.rent) else None,
                "res": r.resultado,
            }
        )
    html = TEMPLATE.read_text(encoding="utf-8")
    html = html.replace("__DATA__", json.dumps(ops, ensure_ascii=False))
    html = html.replace("__UPDATED__", datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M"))
    (DIST / "index.html").write_text(html, encoding="utf-8")
    return ops


def write_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    PURPLE, PANEL, PERSIMON, WHITE, GREY = "0D0A14", "1A1425", "FF5A36", "FFFFFF", "9B93AD"
    MONEY = "#,##0.00;(#,##0.00)"
    wb = Workbook()

    ws = wb.create_sheet("Operaciones")
    heads = [
        "Fecha Cierre", "Símbolo", "Lado", "Tipo", "Promedios",
        "Capital Desplegado ($)", "PnL Bruto ($)", "Comisiones ($)", "Funding ($)",
        "PnL Neto ($)", "Rentabilidad %", "Resultado",
    ]
    ws.append(heads)
    for _, r in df.iterrows():
        ws.append([
            r.close, r.symbol, "Short" if r.side == "short" else "Long", r.tipo,
            int(r.promedios), round(r.capital, 2), round(r.pnl, 2), round(r.fee, 2),
            round(r.funding, 2), round(r.net, 2),
            (float(r.rent) if pd.notna(r.rent) else None), r.resultado,
        ])
    for i in range(2, len(df) + 2):
        ws[f"A{i}"].number_format = "DD/MM/YYYY HH:MM"
        for c in "FGHIJ":
            ws[f"{c}{i}"].number_format = MONEY
        ws[f"K{i}"].number_format = "0.00%"
    for j, _ in enumerate(heads, 1):
        c = ws.cell(1, j)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=PURPLE)
        c.alignment = Alignment(horizontal="center")
    for j, w in enumerate([17, 11, 7, 8, 10, 20, 13, 13, 11, 13, 13, 22], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    wd = wb.create_sheet("Resumen Diario")
    wd.append(["Fecha", "Nº Operaciones", "PnL Neto Día ($)", "Rentabilidad Media Día", "Balance Acumulado ($)"])
    daily = df.assign(fecha=df.close.dt.date).groupby("fecha").agg(
        n=("net", "size"), net=("net", "sum"), rent=("rent", "mean")
    ).reset_index()
    cum = 0.0
    for k, r in enumerate(daily.itertuples(), start=2):
        cum += r.net
        wd.cell(k, 1, pd.Timestamp(r.fecha)).number_format = "DD/MM/YYYY"
        wd.cell(k, 2, int(r.n))
        wd.cell(k, 3, round(r.net, 2)).number_format = MONEY
        if pd.notna(r.rent):
            wd.cell(k, 4, float(r.rent)).number_format = "0.00%"
        wd.cell(k, 5, round(cum, 2)).number_format = MONEY
    for j in range(1, 6):
        c = wd.cell(1, j)
        c.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=PURPLE)
        c.alignment = Alignment(horizontal="center")
    for j, w in enumerate([13, 15, 17, 20, 20], 1):
        wd.column_dimensions[get_column_letter(j)].width = w
    wd.freeze_panes = "A2"

    dsh = wb.create_sheet("Dashboard", 0)
    dsh["B2"] = "DESCENTRALIZADOS — OPERATIVA FUTUROS PIONEX"
    dsh["B2"].font = Font(bold=True, size=16, color=PERSIMON, name="Arial")
    rents = df.rent.dropna()
    metrics = [
        ("% Medio de Ganancia por Operación", float(rents.mean()), "0.00%"),
        ("Balance Total ($)", round(float(df.net.sum()), 2), MONEY),
        ("Media de Ganancia Diaria ($)", round(float(df.net.sum()) / daily.shape[0], 2), MONEY),
    ]
    for j, (lab, val, fmt) in enumerate(metrics):
        col = 2 + j * 2
        c1 = dsh.cell(4, col, lab)
        c1.font = Font(bold=True, size=10, color=GREY, name="Arial")
        c2 = dsh.cell(5, col, val)
        c2.font = Font(bold=True, size=18, color=WHITE, name="Arial")
        c2.number_format = fmt
        for rr in (4, 5):
            dsh.cell(rr, col).fill = PatternFill("solid", start_color=PANEL)
        dsh.column_dimensions[get_column_letter(col)].width = 32
        dsh.column_dimensions[get_column_letter(col + 1)].width = 3
    extra = [
        ("B", "Total operaciones", len(df), None),
        ("D", "Días operados", int(daily.shape[0]), None),
        ("F", "Win rate", float((df.net > 0).mean()), "0.0%"),
    ]
    for colL, lab, val, fmt in extra:
        dsh[f"{colL}7"] = lab
        dsh[f"{colL}7"].font = Font(bold=True, size=10, color=GREY, name="Arial")
        dsh[f"{colL}8"] = val
        dsh[f"{colL}8"].font = Font(bold=True, size=14, color=WHITE, name="Arial")
        dsh[f"{colL}8"].fill = PatternFill("solid", start_color=PANEL)
        if fmt:
            dsh[f"{colL}8"].number_format = fmt
    for row in dsh.iter_rows(min_row=1, max_row=10, min_col=1, max_col=8):
        for c in row:
            if not c.fill.start_color.rgb or c.fill.start_color.rgb == "00000000":
                c.fill = PatternFill("solid", start_color=PURPLE)

    del wb["Sheet"]
    wb.save(DIST / "operativa-futuros-pionex.xlsx")


def main() -> int:
    required = ["position_futures.csv", "raw-trading-details.csv", "others.csv"]
    missing = [f for f in required if not (DATA / f).exists()]
    if missing:
        print(f"ERROR: faltan archivos en data/: {missing}")
        return 1
    pos, raw, risk = load_inputs()
    df = finalize(load_manual_ops(analyze(pos, raw, risk)))
    DIST.mkdir(parents=True, exist_ok=True)
    ops = write_html(df)
    write_excel(df)
    total = sum(o["n"] for o in ops)
    dias = len({o["d"][:10] for o in ops})
    print(f"OK: {len(ops)} operaciones, {dias} días, balance {total:,.2f} $")
    return 0


if __name__ == "__main__":
    sys.exit(main())
